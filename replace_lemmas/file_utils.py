import os
import shutil
import copy
import xml.etree.ElementTree as ET

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

from openpyxl import load_workbook


def parse_xml_entries(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()
    id_to_entry = {}
    id_to_info = {}
    for entry in root.findall('entry'):
        xml_id = entry.get('{http://www.w3.org/XML/1998/namespace}id') or entry.get('xml:id')
        if not xml_id:
            continue
        orth_elem = entry.find("./form[@type='reconstructed']/orth")
        lemma = orth_elem.text.strip() if orth_elem is not None and orth_elem.text else ""
        pb = entry.find('pb')
        page = pb.get('n') if pb is not None else "?"
        try:
            page = int(page)
        except:
            page = page
        id_to_entry[xml_id] = entry
        id_to_info[xml_id] = (lemma, page)
    return root, id_to_entry, id_to_info


def read_excel_manual(excel_path, sheet_name, col_lemma, col_page, col_volume=None, skip_rows=1):
    ext = os.path.splitext(excel_path)[1].lower()
    data = {}

    if ext == '.xls':
        if not HAS_XLRD:
            raise ImportError("Для чтения .xls установите xlrd: pip install xlrd")
        wb = xlrd.open_workbook(excel_path)
        ws = wb.sheet_by_name(sheet_name)
        header_row = skip_rows - 1
        if header_row >= ws.nrows:
            raise ValueError("Лист пуст или skip_rows больше числа строк")
        headers = {}
        for col in range(ws.ncols):
            val = ws.cell_value(header_row, col)
            if val:
                headers[str(val).strip()] = col

        def get_col_idx(name):
            if name in headers:
                return headers[name]
            try:
                return int(name) - 1
            except:
                raise ValueError(f"Столбец '{name}' не найден в заголовках Excel")

        idx_lemma = get_col_idx(col_lemma)
        idx_page = get_col_idx(col_page)
        idx_volume = get_col_idx(col_volume) if col_volume else None

        for row in range(skip_rows, ws.nrows):
            vals = [ws.cell_value(row, c) for c in range(ws.ncols)]
            if len(vals) <= max(idx_lemma, idx_page):
                continue
            lemma = str(vals[idx_lemma]).strip() if vals[idx_lemma] != '' else ""
            page_raw = vals[idx_page]
            try:
                page = int(page_raw)
            except:
                continue
            if not lemma:
                continue
            volume = None
            if idx_volume is not None and len(vals) > idx_volume:
                vol_raw = vals[idx_volume]
                if vol_raw != '':
                    try:
                        volume = int(vol_raw)
                    except:
                        volume = str(vol_raw).strip()
            vol_key = volume if volume is not None else "all"
            data.setdefault(vol_key, {}).setdefault(page, []).append(lemma)

    else:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        header_row = skip_rows
        headers = {}
        for cell in ws[header_row]:
            if cell.value:
                headers[cell.value.strip()] = cell.column - 1

        def get_col_idx(name):
            if name in headers:
                return headers[name]
            try:
                return int(name) - 1
            except:
                raise ValueError(f"Столбец '{name}' не найден в заголовках Excel")

        idx_lemma = get_col_idx(col_lemma)
        idx_page = get_col_idx(col_page)
        idx_volume = get_col_idx(col_volume) if col_volume else None

        for row in ws.iter_rows(min_row=skip_rows + 1, values_only=True):
            if len(row) <= max(idx_lemma, idx_page):
                continue
            lemma = str(row[idx_lemma]).strip() if row[idx_lemma] is not None else ""
            page_raw = row[idx_page]
            try:
                page = int(page_raw)
            except:
                continue
            if not lemma:
                continue
            volume = None
            if idx_volume is not None and len(row) > idx_volume:
                vol_raw = row[idx_volume]
                if vol_raw is not None:
                    try:
                        volume = int(vol_raw)
                    except:
                        volume = str(vol_raw).strip()
            vol_key = volume if volume is not None else "all"
            data.setdefault(vol_key, {}).setdefault(page, []).append(lemma)

        wb.close()
    return data


def backup_files(xml_path, entries_dir):
    backup_xml = xml_path + '.bak'
    backup_entries = entries_dir + '.bak'
    shutil.copy2(xml_path, backup_xml)
    if os.path.exists(backup_entries):
        shutil.rmtree(backup_entries)
    shutil.copytree(entries_dir, backup_entries)
    return backup_xml, backup_entries


def restore_backup(xml_path, entries_dir, backup_xml, backup_entries):
    shutil.copy2(backup_xml, xml_path)
    if os.path.exists(entries_dir):
        shutil.rmtree(entries_dir)
    shutil.copytree(backup_entries, entries_dir)


def save_xml_and_entries(root, xml_path, entries_dir):
    ET.indent(root, space="  ")
    tree = ET.ElementTree(root)
    tree.write(xml_path, encoding='utf-8', xml_declaration=True)

    for fname in os.listdir(entries_dir):
        if fname.endswith('.xml'):
            os.remove(os.path.join(entries_dir, fname))

    for entry in root.findall('entry'):
        xml_id = entry.get('{http://www.w3.org/XML/1998/namespace}id') or entry.get('xml:id')

        single_root = ET.Element('Entry')
        single_root.set('xml:id', xml_id)

        for child in entry:
            single_root.append(copy.deepcopy(child))

        ET.indent(single_root, space="  ")
        xml_str = ET.tostring(single_root, encoding='unicode')
        fpath = os.path.join(entries_dir, f"{xml_id}.xml")
        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(xml_str)