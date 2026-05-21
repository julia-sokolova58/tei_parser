import os
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from replace_lemmas.file_utils import (
    parse_xml_entries, read_excel_manual, backup_files,
    restore_backup, save_xml_and_entries
)
from replace_lemmas.core import replace_lemmas, apply_doubtful_decisions, apply_mismatch_decisions
from replace_lemmas.dialogs import DoubtfulDialog, MismatchDialog


def find_available_volumes(output_dir="output"):
    if not os.path.isdir(output_dir):
        return []
    volumes = []
    for item in os.listdir(output_dir):
        item_path = os.path.join(output_dir, item)
        if os.path.isdir(item_path):
            xml_path = os.path.join(item_path, f"{item}.xml")
            entries_path = os.path.join(item_path, f"{item}_entries")
            if os.path.isfile(xml_path) and os.path.isdir(entries_path):
                volumes.append(item)
    return sorted(volumes)


def extract_volume_number(volume_name):
    import re
    match = re.search(r'(\d+)$', volume_name)
    return match.group(1) if match else volume_name


class RedirectText:
    def __init__(self, text_widget):
        self.text_widget = text_widget
    def write(self, string):
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)
    def flush(self):
        pass


class ReplaceApp:
    def __init__(self, root):
        self.root = root
        root.title("Замена лемм с ручной проверкой")
        root.geometry("750x720")

        self.output_dir = "output"
        self.volume_var = tk.StringVar(value="")
        self.excel_var = tk.StringVar(value="тома_1-37_леммы.xls")
        self.sheet_var = tk.StringVar()
        self.col_lemma_var = tk.StringVar(value="Корень")
        self.col_page_var = tk.StringVar(value="Страница")
        self.skip_rows_var = tk.StringVar(value="2")
        self.sim_threshold_var = tk.StringVar(value="0.3")
        self.offset_var = tk.StringVar(value="0")

        frm_volume = tk.Frame(root)
        frm_volume.pack(pady=5, padx=10, fill=tk.X)
        tk.Label(frm_volume, text="Выберите том из папки output:").pack(anchor=tk.W)
        self.combo_volume = ttk.Combobox(frm_volume, textvariable=self.volume_var, width=40, state='readonly')
        self.combo_volume.pack(side=tk.LEFT, padx=5)
        tk.Button(frm_volume, text="Обновить список томов", command=self.refresh_volumes).pack(side=tk.LEFT, padx=5)

        frm_excel = tk.Frame(root)
        frm_excel.pack(pady=5, padx=10, fill=tk.X)
        tk.Label(frm_excel, text="Excel-файл ручной росписи:").pack(anchor=tk.W)
        self.entry_excel = tk.Entry(frm_excel, textvariable=self.excel_var, width=50)
        self.entry_excel.pack(side=tk.LEFT, expand=True, fill=tk.X)
        tk.Button(frm_excel, text="Обзор...", command=self.browse_excel).pack(side=tk.RIGHT, padx=5)

        frm_sheet = tk.Frame(root)
        frm_sheet.pack(pady=5, padx=10, fill=tk.X)
        tk.Label(frm_sheet, text="Лист Excel (авто-выбор по тому):").pack(side=tk.LEFT)
        self.combo_sheet = ttk.Combobox(frm_sheet, textvariable=self.sheet_var, width=30, state='readonly')
        self.combo_sheet.pack(side=tk.LEFT, padx=5)
        tk.Button(frm_sheet, text="Обновить листы", command=self.refresh_sheets).pack(side=tk.LEFT, padx=5)

        frm_cols = tk.Frame(root)
        frm_cols.pack(pady=5, padx=10, fill=tk.X)
        tk.Label(frm_cols, text="Столбцы в Excel (имена или номера):").pack(anchor=tk.W)
        grid = tk.Frame(frm_cols)
        grid.pack(anchor=tk.W)
        tk.Label(grid, text="Лемма:").grid(row=0, column=0, sticky=tk.W, padx=2)
        self.entry_col_lemma = tk.Entry(grid, textvariable=self.col_lemma_var, width=15)
        self.entry_col_lemma.grid(row=0, column=1, padx=2)
        tk.Label(grid, text="Страница:").grid(row=0, column=2, sticky=tk.W, padx=2)
        self.entry_col_page = tk.Entry(grid, textvariable=self.col_page_var, width=15)
        self.entry_col_page.grid(row=0, column=3, padx=2)

        frm_sim = tk.Frame(root)
        frm_sim.pack(pady=5, padx=10, fill=tk.X)
        tk.Label(frm_sim, text="Минимальное сходство (0..1):").pack(side=tk.LEFT)
        self.entry_sim = tk.Entry(frm_sim, textvariable=self.sim_threshold_var, width=5)
        self.entry_sim.pack(side=tk.LEFT, padx=5)

        frm_skip = tk.Frame(root)
        frm_skip.pack(pady=5, padx=10, fill=tk.X)
        tk.Label(frm_skip, text="Пропустить строк в начале (включая заголовки):").pack(side=tk.LEFT)
        self.entry_skip_rows = tk.Entry(frm_skip, textvariable=self.skip_rows_var, width=5)
        self.entry_skip_rows.pack(side=tk.LEFT, padx=5)

        frm_offset = tk.Frame(root)
        frm_offset.pack(pady=5, padx=10, fill=tk.X)
        tk.Label(frm_offset, text="Смещение страниц (±):").pack(side=tk.LEFT)
        self.entry_offset = tk.Entry(frm_offset, textvariable=self.offset_var, width=5)
        self.entry_offset.pack(side=tk.LEFT, padx=5)
        tk.Label(frm_offset, text="(добавляется к странице из XML при сравнении с Excel)").pack(side=tk.LEFT)

        self.btn_replace = tk.Button(root, text="Заменить леммы", command=self.start_replace, bg="#4CAF50", fg="white", font=("Arial", 10, "bold"))
        self.btn_replace.pack(pady=10)

        self.log_widget = scrolledtext.ScrolledText(root, height=20, wrap=tk.WORD)
        self.log_widget.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)

        self.redirect = RedirectText(self.log_widget)
        sys.stdout = self.redirect
        sys.stderr = self.redirect

        self.refresh_volumes()
        self.root.after(200, self.refresh_sheets)

    def refresh_volumes(self):
        volumes = find_available_volumes(self.output_dir)
        self.combo_volume['values'] = volumes
        if volumes:
            self.volume_var.set(volumes[0])
            self.refresh_sheets()

    def browse_excel(self):
        file = filedialog.askopenfilename(
            title="Выберите Excel-файл с росписью",
            filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
        )
        if file:
            self.excel_var.set(file)
            self.refresh_sheets()

    def refresh_sheets(self):
        excel = self.excel_var.get()
        if not excel or not os.path.isfile(excel):
            return
        try:
            from openpyxl import load_workbook
            try:
                import xlrd
                HAS_XLRD = True
            except ImportError:
                HAS_XLRD = False

            ext = os.path.splitext(excel)[1].lower()
            sheets = []
            if ext == '.xls':
                if not HAS_XLRD:
                    raise ImportError("Установите xlrd: pip install xlrd")
                wb = xlrd.open_workbook(excel)
                sheets = wb.sheet_names()
            else:
                wb = load_workbook(excel, read_only=True, data_only=True)
                sheets = wb.sheetnames
                wb.close()
            self.combo_sheet['values'] = sheets

            volume = self.volume_var.get()
            if volume:
                volume_num = extract_volume_number(volume)
                if volume_num in sheets:
                    self.sheet_var.set(volume_num)
                elif volume in sheets:
                    self.sheet_var.set(volume)
                elif sheets:
                    self.sheet_var.set(sheets[0])
            elif sheets:
                self.sheet_var.set(sheets[0])
        except Exception as e:
            self.log(f"Ошибка чтения Excel: {e}")

    def log(self, message):
        self.log_widget.insert(tk.END, message + '\n')
        self.log_widget.see(tk.END)

    def start_replace(self):
        volume = self.volume_var.get()
        if not volume:
            messagebox.showerror("Ошибка", "Выберите том из списка")
            return

        volume_path = os.path.join(self.output_dir, volume)
        xml_path = os.path.join(volume_path, f"{volume}.xml")
        entries_dir = os.path.join(volume_path, f"{volume}_entries")

        if not os.path.isfile(xml_path):
            messagebox.showerror("Ошибка", f"XML не найден: {xml_path}")
            return
        if not os.path.isdir(entries_dir):
            messagebox.showerror("Ошибка", f"Папка entries не найдена: {entries_dir}")
            return

        excel = self.excel_var.get().strip()
        if not excel or not os.path.isfile(excel):
            messagebox.showerror("Ошибка", "Укажите корректный Excel-файл")
            return

        sheet = self.sheet_var.get()
        if not sheet:
            messagebox.showerror("Ошибка", "Выберите лист Excel")
            return

        col_lemma = self.col_lemma_var.get().strip()
        col_page = self.col_page_var.get().strip()

        if not col_lemma or not col_page:
            messagebox.showerror("Ошибка", "Укажите названия или номера столбцов леммы и страницы")
            return

        try:
            skip_rows = int(self.skip_rows_var.get())
            if skip_rows < 1:
                raise ValueError
            sim_threshold = float(self.sim_threshold_var.get().replace(',', '.'))
            if not (0 < sim_threshold <= 1):
                raise ValueError
            offset = int(self.offset_var.get().strip())
        except:
            messagebox.showerror("Ошибка", "Проверьте числовые значения.")
            return

        if not messagebox.askyesno("Подтверждение",
            f"Том: {volume}\n"
            f"XML: {xml_path}\n"
            f"Excel: {os.path.basename(excel)}\n"
            f"Лист: {sheet}\n\n"
            "Будут изменены XML и файлы entries. Продолжить?"):
            return

        try:
            backup_xml, backup_entries = backup_files(xml_path, entries_dir)
        except Exception as e:
            self.log(f"Ошибка создания резервной копии: {e}")
            return

        self.btn_replace.config(state=tk.DISABLED, text="Идёт замена...")
        self.log_widget.delete(1.0, tk.END)
        self.log(f"Том: {volume}")
        self.log(f"XML: {xml_path}")
        self.log(f"Excel: {excel}")
        self.log(f"Лист: {sheet}")
        self.log(f"Столбец леммы: {col_lemma}")
        self.log(f"Столбец страницы: {col_page}")
        self.log(f"Смещение страниц: {offset}")
        self.log("-" * 50)

        def worker():
            try:
                root, id_to_entry, id_to_info = parse_xml_entries(xml_path)
                manual = read_excel_manual(excel, sheet, col_lemma, col_page, None, skip_rows)
                root, auto_log, doubtful, mismatch = replace_lemmas(
                    root, id_to_entry, id_to_info, manual, volume=extract_volume_number(volume),
                    sim_threshold=sim_threshold, offset=offset
                )
                for line in auto_log:
                    self.log(line)

                if doubtful:
                    self.log(f"\nНайдено {len(doubtful)} сомнительных пар. Открываю ручную проверку...")
                    dialog_res = []

                    def run_doubtful():
                        dlg = DoubtfulDialog(self.root, doubtful)
                        self.root.wait_window(dlg)
                        dialog_res.append(dlg.result)
                        if dlg.result:
                            dialog_res.append(dlg.decisions)

                    self.root.after(0, run_doubtful)
                    import time
                    while not dialog_res:
                        time.sleep(0.1)

                    if dialog_res[0]:
                        decisions = dialog_res[1]
                        dlog = apply_doubtful_decisions(root, id_to_entry, id_to_info, doubtful, decisions)
                        for line in dlog:
                            self.log(line)
                    else:
                        self.log("Ручная проверка отменена. Изменения отменены.")
                        restore_backup(xml_path, entries_dir, backup_xml, backup_entries)
                        self.btn_replace.config(state=tk.NORMAL, text="Заменить леммы")
                        return

                if mismatch:
                    self.log(f"\nНайдено {len(mismatch)} страниц с разным количеством. Открываю сопоставление...")
                    mm_res = []

                    def run_mismatch():
                        dlg = MismatchDialog(self.root, mismatch)
                        self.root.wait_window(dlg)
                        mm_res.append(dlg.result)
                        if dlg.result:
                            mm_res.append((dlg.lemma_decisions, dlg.page_decisions))

                    self.root.after(0, run_mismatch)
                    import time
                    while not mm_res:
                        time.sleep(0.1)

                    if mm_res[0]:
                        lemma_dec, page_dec = mm_res[1]
                        mlog = apply_mismatch_decisions(root, id_to_entry, mismatch, lemma_dec, page_dec)
                        for line in mlog:
                            self.log(line)
                    else:
                        self.log("Сопоставление отменено. Изменения отменены.")
                        restore_backup(xml_path, entries_dir, backup_xml, backup_entries)
                        self.btn_replace.config(state=tk.NORMAL, text="Заменить леммы")
                        return

                save_xml_and_entries(root, xml_path, entries_dir)

                self.log("\n" + "=" * 50)
                self.log("Обработка завершена успешно!")
                self.log(f"Общий XML: {xml_path}")
                self.log(f"Постатейные файлы: {entries_dir}")
                self.log("Резервные копии: .bak")
            except Exception as e:
                self.log(f"\nОшибка: {e}")
                import traceback
                self.log(traceback.format_exc())
                restore_backup(xml_path, entries_dir, backup_xml, backup_entries)
                self.log("Изменения отменены (восстановлены резервные копии).")
            finally:
                self.btn_replace.config(state=tk.NORMAL, text="Заменить леммы")

        threading.Thread(target=worker, daemon=True).start()


def main():
    root = tk.Tk()
    app = ReplaceApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()